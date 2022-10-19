import eel
import openpyxl

eel.init('web')

@eel.expose
def entry_excel(i):
    doc = openpyxl.load_workbook("bus.xlsx")
    sheet = doc.active

    cell1 = sheet.cell(row= i, column=2)

    try:
        cell1.value += 1
    except:
        cell1.value = 1

    row = sheet.max_row
    col = sheet.max_column
    total = 0

    for a in range(1, row+1):
        cell2 = sheet.cell(row=a, column=2)
        try:
            total += cell2.value
        except:
            pass
    
    if total < 6:
        doc.save("bus.xlsx")
        doc.close
        return 1
    else:
        return 0
    

eel.start('index.html')
